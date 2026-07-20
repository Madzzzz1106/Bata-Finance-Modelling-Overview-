import openpyxl

wb = openpyxl.load_workbook("Campus Activewe.xlsx", data_only=False)
ws = wb["Profit & Loss"]
# Print first 10 rows and first 12 columns formulas
print("--- Profit & Loss Formulas ---")
for r in range(1, 11):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 13)]
    print(f"Row {r}: {row_vals}")

print("\n--- Data Sheet dimensions and sample ---")
ws_data = wb["Data Sheet"]
print(f"Data Sheet dimensions: {ws_data.dimensions}")
for r in range(1, 11):
    row_vals = [ws_data.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")
