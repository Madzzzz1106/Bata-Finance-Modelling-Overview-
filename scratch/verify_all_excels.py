import openpyxl
import glob

for f in sorted(glob.glob("*.xlsx")):
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb["Data Sheet"]
    print(f"\n===== File: {f} =====")
    print(f"Row 15 label: {ws.cell(row=15, column=1).value}")
    print(f"Row 16 label: {ws.cell(row=16, column=1).value}")
    print(f"Row 17 label: {ws.cell(row=17, column=1).value}")
    
    print(f"Row 55 label: {ws.cell(row=55, column=1).value}")
    print(f"Row 56 label: {ws.cell(row=56, column=1).value}")
    print(f"Row 57 label: {ws.cell(row=57, column=1).value}")
    
    print(f"Row 80 label: {ws.cell(row=80, column=1).value}")
    print(f"Row 81 label: {ws.cell(row=81, column=1).value}")
    print(f"Row 82 label: {ws.cell(row=82, column=1).value}")
