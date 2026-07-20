import pandas as pd
import fitz
import glob
import os

def check_1_fy20():
    print("=== CHECK 1: FY20 Comparative Columns ===")
    for company in ["Bata", "Relaxo", "Metro"]:
        filename = f"{company}_FY21.pdf"
        doc = fitz.open(filename)
        # Check text in Balance Sheet pages
        # Bata page 199, Metro page 99, Relaxo page 76
        p = 199 if company == "Bata" else (99 if company == "Metro" else 76)
        text = doc[p-1].get_text("text")
        print(f"\n{company} FY21 PDF Page {p} header lines:")
        lines = [l.strip() for l in text.split("\n") if l.strip()][:15]
        print(lines)

def check_2_fy26():
    print("\n=== CHECK 2: Screener FY26 Data ===")
    for f in ["Campus Activewe.xlsx", "Khadim India.xlsx", "Liberty Shoes.xlsx"]:
        wb = fitz.open() # just to keep it simple, use openpyxl
        import openpyxl
        wb_xl = openpyxl.load_workbook(f, data_only=False)
        ws = wb_xl["Data Sheet"]
        # Col K (column 11) is the last column
        col_k_header = ws.cell(row=16, column=11).value
        # Check if there is any other column
        print(f"{f}: Col 11 header = {col_k_header}")
        # Look at the formula for trailing on Profit & Loss
        ws_pl = wb_xl["Profit & Loss"]
        print(f"  P&L Row 3 (Headers): {[ws_pl.cell(row=3, column=c).value for c in range(1, 14)]}")
        print(f"  P&L Row 4 (Sales): {[ws_pl.cell(row=4, column=c).value for c in range(1, 14)]}")

def check_3_validation():
    print("\n=== CHECK 3: Balance Sheet Equation (Total Assets vs Equity + Liabilities) ===")
    # Bata FY25
    df_bata = pd.read_csv("extracted/Bata_FY25_balance_sheet.csv")
    ta_bata = float(df_bata[df_bata["Particulars"].str.lower().str.contains("total assets")]["CurrentYear"].values[0])
    te_bata = float(df_bata[df_bata["Particulars"].str.lower().str.contains("total equity") & ~df_bata["Particulars"].str.lower().str.contains("and liabilities")]["CurrentYear"].values[0])
    tl_bata = float(df_bata[df_bata["Particulars"].str.lower().str.contains("total liabilities") & ~df_bata["Particulars"].str.lower().str.contains("equity")]["CurrentYear"].values[0])
    tel_bata = float(df_bata[df_bata["Particulars"].str.lower().str.contains("total equity and liabilities")]["CurrentYear"].values[0])
    print(f"Bata FY25: Total Assets = {ta_bata}, Total Equity = {te_bata}, Total Liabilities = {tl_bata}, Sum = {te_bata+tl_bata}, Total Equity and Liabilities = {tel_bata}")
    
    # Relaxo FY25
    df_relaxo = pd.read_csv("extracted/Relaxo_FY25_balance_sheet.csv")
    ta_relaxo = float(df_relaxo[df_relaxo["Particulars"].str.strip().str.lower() == "total assets"]["CurrentYear"].values[0])
    # Relaxo doesn't have "Total Equity" row, so sum Equity Share Capital and Other Equity
    eq_share_cap = float(df_relaxo[df_relaxo["Particulars"].str.lower().str.contains("equity share capital")]["CurrentYear"].values[0])
    other_eq = float(df_relaxo[df_relaxo["Particulars"].str.lower().str.contains("other equity")]["CurrentYear"].values[0])
    te_relaxo = eq_share_cap + other_eq
    
    # Sum up all liabilities
    # Total Equity and Liabilities is 2762.47
    tel_relaxo = float(df_relaxo[df_relaxo["Particulars"].str.strip().str.lower() == "total equity and liabilities"]["CurrentYear"].values[0])
    tl_relaxo = tel_relaxo - te_relaxo
    print(f"Relaxo FY25: Total Assets = {ta_relaxo}, Total Equity (calculated) = {te_relaxo}, Total Liabilities (calculated) = {tl_relaxo:.2f}, Sum = {te_relaxo+tl_relaxo:.2f}, Total Equity and Liabilities = {tel_relaxo}")
    
    # Metro FY25
    df_metro = pd.read_csv("extracted/Metro_FY25_balance_sheet.csv")
    # Metro column names are CurrentYear, PreviousYear
    ta_metro = float(df_metro[df_metro["Particulars"].str.strip().str.lower() == "total assets"]["CurrentYear"].values[0])
    # Metro has "Total equity" row
    te_metro_rows = df_metro[df_metro["Particulars"].str.strip().str.lower() == "total equity"]["CurrentYear"].values
    if len(te_metro_rows) > 0:
        te_metro = float(te_metro_rows[0])
    else:
        # fallback
        te_metro = float(df_metro[df_metro["Particulars"].str.lower().str.contains("equity share capital")]["CurrentYear"].values[0]) + float(df_metro[df_metro["Particulars"].str.lower().str.contains("other equity")]["CurrentYear"].values[0])
        
    tel_metro = float(df_metro[df_metro["Particulars"].str.lower().str.contains("total equity and liabilities")]["CurrentYear"].values[0])
    tl_metro = tel_metro - te_metro
    print(f"Metro FY25: Total Assets = {ta_metro}, Total Equity = {te_metro}, Total Liabilities (calculated) = {tl_metro:.2f}, Sum = {te_metro+tl_metro:.2f}, Total Equity and Liabilities = {tel_metro}")

def check_4_rowcounts():
    print("\n=== CHECK 4: Detailed Row Count Breakdown ===")
    companies = ["Bata", "Relaxo", "Metro", "Campus Activewear", "Khadim India", "Liberty Shoes"]
    for c in companies:
        print(f"\nCompany: {c}")
        # CSV files
        for stmt in ["balance_sheet", "pnl", "cashflow"]:
            path = glob.glob(f"extracted/{c}_*_{stmt}.csv") + glob.glob(f"extracted/{c}_{stmt}.csv")
            if path:
                # Count total rows across all matching files
                total_rows = 0
                for p in path:
                    df = pd.read_csv(p)
                    total_rows += len(df)
                print(f"  {stmt.upper()}: {total_rows} rows (in {len(path)} files)")
            else:
                print(f"  {stmt.upper()}: 0 rows")
        # Text files
        for txt_type in ["mda", "related_party", "contingent_liabilities", "auditor_report"]:
            path = glob.glob(f"extracted/{c}_*_{txt_type}.txt")
            if path:
                total_lines = 0
                for p in path:
                    with open(p, "r") as f:
                        total_lines += len(f.readlines())
                print(f"  {txt_type.upper()} text: {total_lines} lines (in {len(path)} files)")
            else:
                print(f"  {txt_type.upper()} text: 0 lines")

if __name__ == "__main__":
    check_1_fy20()
    check_2_fy26()
    check_3_validation()
    check_4_rowcounts()
