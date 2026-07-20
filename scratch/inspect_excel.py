import openpyxl

for fname in ["Campus Activewe.xlsx", "Khadim India.xlsx", "Liberty Shoes.xlsx"]:
    print(f"=== {fname} ===")
    wb = openpyxl.load_workbook(fname, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"  Sheet: {sname}")
        row_names = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                row_names.append((r, str(val).strip()))
        print(f"    Rows with labels: {len(row_names)}")
        for r, name in row_names[:45]:
            print(f"      {r:03d}: {name}")
        if len(row_names) > 45:
            print("      ...")
            for r, name in row_names[-15:]:
                print(f"      {r:03d}: {name}")
