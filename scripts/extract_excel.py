import openpyxl
import pandas as pd
import os
import glob
from datetime import datetime

# Map Excel file name to desired CompanyName
COMPANY_MAP = {
    "Campus Activewe.xlsx": "Campus Activewear",
    "Khadim India.xlsx": "Khadim India",
    "Liberty Shoes.xlsx": "Liberty Shoes"
}

def parse_section(ws, start_row, end_row, header_row):
    # Find columns that have valid dates in the header row
    cols = []
    years = []
    
    # Check columns from 3 (C) onwards
    c = 3
    while True:
        val = ws.cell(row=header_row, column=c).value
        if val is None:
            break
        # Sometimes it is a datetime, sometimes a string/number
        if isinstance(val, datetime):
            year_str = str(val.year)
        else:
            # try to parse or convert to string
            year_str = str(val)[:4]
        
        cols.append(c)
        years.append(year_str)
        c += 1
        
    rows_data = []
    for r in range(start_row, end_row + 1):
        metric = ws.cell(row=r, column=1).value
        if not metric:
            continue
            
        row_vals = []
        for col_idx in cols:
            val = ws.cell(row=r, column=col_idx).value
            # Clean number value
            if val is None or val == "":
                row_vals.append("")
            else:
                try:
                    row_vals.append(float(val))
                except ValueError:
                    row_vals.append(str(val))
                    
        rows_data.append([metric] + row_vals)
        
    df = pd.DataFrame(rows_data, columns=["Particulars"] + years)
    return df

def extract_excel_data():
    for excel_file, company_name in COMPANY_MAP.items():
        if not os.path.exists(excel_file):
            print(f"Warning: Excel file {excel_file} not found. Skipping.")
            continue
            
        print(f"Extracting Excel: {excel_file} -> {company_name}")
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Data Sheet"]
        
        # 1. P&L
        df_pnl = parse_section(ws, start_row=17, end_row=30, header_row=16)
        pnl_path = os.path.join("extracted", f"{company_name}_pnl.csv")
        df_pnl.to_csv(pnl_path, index=False)
        print(f"  Saved P&L to {pnl_path} ({len(df_pnl)} rows)")
        
        # 2. Balance Sheet
        df_bs = parse_section(ws, start_row=57, end_row=69, header_row=56)
        bs_path = os.path.join("extracted", f"{company_name}_balance_sheet.csv")
        df_bs.to_csv(bs_path, index=False)
        print(f"  Saved Balance Sheet to {bs_path} ({len(df_bs)} rows)")
        
        # 3. Cash Flow
        df_cf = parse_section(ws, start_row=82, end_row=85, header_row=81)
        cf_path = os.path.join("extracted", f"{company_name}_cashflow.csv")
        df_cf.to_csv(cf_path, index=False)
        print(f"  Saved Cash Flow to {cf_path} ({len(df_cf)} rows)")

if __name__ == "__main__":
    extract_excel_data()
